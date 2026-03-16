"""
util/parse_financials.py

Parses a Microsoft quarterly financial statements XLSX and returns a flat dict
of all CompanyState Financials + Segments fields (USD millions, floats).

── Confirmed file structure (from FinancialStatementFY23Q1.xlsx inspection) ──
Sheet names (exact):
  "Income Statements"      → revenue, costs, margins, operating/net income
  "Balance Sheets"         → cash and cash equivalents
  "Cash Flows"             → capex (Additions to property and equipment)
  "Segment Revenue & OI"   → productivity, intelligent cloud, personal computing

Column layout (all sheets):
  col A (index 0) = row label
  col B (index 1) = current quarter value   ← this is what we want
  col C (index 2) = footnote marker (sometimes)
  col D (index 3) = prior year comparison value

Value types: int or float (not strings). openpyxl data_only=True reads them
directly — no string parsing needed.

Note on gross_margin: the XLSX stores gross margin as an absolute dollar value
(e.g. 34670), not a percentage. We compute the percentage ourselves:
  gross_margin_pct = gross_margin_dollars / total_revenue * 100
"""

from __future__ import annotations
from pathlib import Path
import openpyxl


def _get_sheet(wb: openpyxl.Workbook, name: str):
    """Return sheet by exact name, or None if not found."""
    return wb[name] if name in wb.sheetnames else None


def _find_value(sheet, label_pattern: str, value_col_index: int = 1) -> float:
    """
    Scan every row in sheet for a cell (col A) whose stripped string contains
    label_pattern (case-insensitive). Return the float at value_col_index
    (0-indexed), or 0.0 if not found.

    Uses substring match with strip() to handle leading spaces in row labels
    (e.g. "   Cash and cash equivalents" → matches "cash and cash equivalents").
    """
    if sheet is None:
        return 0.0
    needle = label_pattern.lower()
    for row in sheet.iter_rows(values_only=True):
        label = row[0]
        if label is None:
            continue
        # Normalize: strip whitespace and newlines, lowercase
        normalized = str(label).replace("\n", " ").strip().lower()
        if needle in normalized:
            try:
                val = row[value_col_index]
                return float(val) if val is not None else 0.0
            except (TypeError, ValueError, IndexError):
                return 0.0
    return 0.0


def parse_financials(path: Path) -> dict:
    """
    Parse a Microsoft quarterly financial statements XLSX.
    Returns a flat dict of CompanyState Financials + Segments fields.
    All monetary values are USD millions (floats).
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    income  = _get_sheet(wb, "Income Statements")
    balance = _get_sheet(wb, "Balance Sheets")
    cashflow= _get_sheet(wb, "Cash Flows")
    segment = _get_sheet(wb, "Segment Revenue & OI")

    # ── Income Statement ──────────────────────────────────────────────────────
    revenue       = _find_value(income, "total revenue")
    cost_of_rev   = _find_value(income, "total cost of revenue")
    gross_margin_abs = _find_value(income, "gross margin")
    operating_inc = _find_value(income, "operating income")
    net_income    = _find_value(income, "net income")
    rd_spending   = _find_value(income, "research and development")
    sales_mktg    = _find_value(income, "sales and marketing")
    gen_admin     = _find_value(income, "general and administrative")

    # Compute gross margin as a percentage (XLSX stores absolute value)
    gross_margin_pct = (gross_margin_abs / revenue * 100) if revenue else 0.0

    # Total operating expenses = R&D + Sales & Marketing + G&A
    total_opex = rd_spending + sales_mktg + gen_admin

    # ── Balance Sheet ─────────────────────────────────────────────────────────
    # "Cash and cash equivalents, end of period" is in Cash Flows (more precise)
    # Balance sheet also has it — using Cash Flows end-of-period value
    cash = _find_value(cashflow, "cash and cash equivalents, end of period")
    if cash == 0.0:
        # Fallback to balance sheet
        cash = _find_value(balance, "cash and cash equivalents")

    # ── Cash Flow Statement ───────────────────────────────────────────────────
    # Capex = "Additions to property and equipment" — stored as negative in XLSX
    capex_raw = _find_value(cashflow, "additions to property and equipment")
    capex = abs(capex_raw)   # store as positive number

    # ── Segment Revenue ───────────────────────────────────────────────────────
    # The segment sheet has two "Productivity and Business Processes" rows:
    # row 10 = Revenue, row 15 = Operating Income
    # _find_value() returns the FIRST match → Revenue row (correct)
    productivity_rev = _find_value(segment, "productivity and business processes")
    intelligent_cloud_rev = _find_value(segment, "intelligent cloud")
    personal_computing_rev = _find_value(segment, "more personal computing")

    return {
        # ── Financials ────────────────────────────────────────────────────────
        "revenue":                  revenue,
        "cost_of_revenue":          cost_of_rev,
        "gross_margin":             round(gross_margin_pct, 2),
        "operating_income":         operating_inc,
        "net_income":               net_income,
        "cash_and_equivalents":     cash,
        "total_operating_expenses": total_opex,
        "rd_spending":              rd_spending,
        "sales_marketing_spending": sales_mktg,
        "capex":                    capex,
        # ── Segments ──────────────────────────────────────────────────────────
        "productivity_revenue":        productivity_rev,
        "intelligent_cloud_revenue":   intelligent_cloud_rev,
        "personal_computing_revenue":  personal_computing_rev,
    }